# desercion_app/app/utils/exportaciones.py
import os
import pandas as pd
from fpdf import FPDF
from matplotlib import image as mpimg
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage


def generate_pdf(resultados, items=None):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Título
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(0, 10, "Resultados de Predicción de Abandono Escolar", ln=True, align="C")
    pdf.ln(10)

    # Métricas
    if 'metricas' in resultados:
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 10, "Métricas del Modelo", ln=True)
        pdf.set_font("Arial", '', 11)
        for k, v in resultados['metricas'].items():
            pdf.cell(0, 8, f"{k.title()}: {v}", ln=True)
        pdf.ln(5)

    # Coeficientes del Modelo
    if 'modelo_coeficientes' in resultados:
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 10, "Coeficientes del Modelo", ln=True)
        pdf.set_font("Arial", '', 11)
        for k, v in resultados['modelo_coeficientes'].items():
            pdf.cell(0, 8, f"{k}: {round(v, 4)}", ln=True)
        pdf.ln(5)

    # Tabla de predicciones
    if 'datos_predicciones' in resultados:
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 10, "Predicciones", ln=True)
        pdf.set_font("Arial", '', 10)
        headers = resultados['datos_predicciones'][0].keys()
        for row in resultados['datos_predicciones'][:20]:  # Limitar a 20 por tamaño
            linea = ', '.join(str(row[col]) for col in headers)
            pdf.multi_cell(0, 8, linea)
        pdf.ln(5)

    # Gráficos
    for graf_path in [resultados.get('grafico_roc'), resultados.get('grafico_prec_rec')]:
        if graf_path and os.path.exists(graf_path):
            pdf.image(graf_path, w=160)
            pdf.ln(5)

    # Guardar PDF
    output_path = os.path.join("desercion_app", "app", "static", "resultados_exportados.pdf")
    pdf.output(output_path)
    return output_path


def generate_excel(resultados, items=None):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Métricas"
    
    # Métricas
    ws1.append(["Métrica", "Valor"])
    for k, v in resultados['metricas'].items():
        ws1.append([k, v])

    # Coeficientes
    ws2 = wb.create_sheet("Coeficientes")
    ws2.append(["Variable", "Coeficiente"])
    for var, val in resultados['modelo_coeficientes'].items():
        ws2.append([var, val])

    # Predicciones
    ws3 = wb.create_sheet("Predicciones")
    if resultados['datos_predicciones']:
        headers = resultados['datos_predicciones'][0].keys()
        ws3.append(list(headers))
        for row in resultados['datos_predicciones']:
            ws3.append([row[h] for h in headers])

    # Insertar gráficos
    for path, title in zip(
        [resultados.get('grafico_roc'), resultados.get('grafico_prec_rec')],
        ["Curva ROC", "Precision-Recall"]
    ):
        if path and os.path.exists(path):
            ws = wb.create_sheet(title)
            img = XLImage(path)
            img.width = 640
            img.height = 480
            ws.add_image(img, "A1")

    output_path = os.path.join("desercion_app", "app", "static", "resultados_exportados.xlsx")
    wb.save(output_path)
    return output_path
