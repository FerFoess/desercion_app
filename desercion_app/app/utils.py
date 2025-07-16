from fpdf import FPDF
import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

EXPORT_FOLDER = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data'))
TEMP_FOLDER = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'static', 'temp'))

def generate_pdf(datos, items):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="Reporte de Resultados", ln=True, align='C')
    pdf.ln(10)
    pdf.set_font("Arial", size=10)

    # Tabla resumen
    if 'datos_predicciones' in datos:
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 10, "Tabla de predicciones", ln=True)
        pdf.set_font("Arial", size=10)

        tabla = pd.DataFrame(datos['datos_predicciones'])
        col_names = tabla.columns.tolist()
        pdf.set_fill_color(220, 220, 220)
        for col in col_names:
            pdf.cell(48, 8, col[:15], border=1, fill=True)
        pdf.ln()
        for _, row in tabla.iterrows():
            for val in row:
                pdf.cell(48, 8, str(val)[:15], border=1)
            pdf.ln()
        pdf.ln(5)

    # Métricas
    if 'metricas' in datos:
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 10, "Métricas del modelo", ln=True)
        pdf.set_font("Arial", size=10)
        for key, val in datos['metricas'].items():
            pdf.cell(0, 10, f"{key}: {val}", ln=True)
        pdf.ln(5)

    # Coeficientes
    if 'modelo_coeficientes' in datos:
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 10, "Coeficientes del modelo", ln=True)
        pdf.set_font("Arial", size=10)
        for k, v in datos['modelo_coeficientes'].items():
            pdf.cell(0, 10, f"{k}: {round(v, 4)}", ln=True)
        pdf.ln(5)

    # 🔽 Gráficos del modelo
    for key, titulo in [('grafico_roc', 'Curva ROC'), ('grafico_prec_rec', 'Curva Precisión-Recall')]:
        img_path = datos.get(key)
        if img_path and os.path.isfile(img_path):
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(0, 10, f"Gráfico: {titulo}", ln=True)
            pdf.image(img_path, w=160)
            pdf.ln(10)

    # 🔽 Imágenes exploratorias en static/temp
    if os.path.exists(TEMP_FOLDER):
        imagenes = [f for f in os.listdir(TEMP_FOLDER) if f.endswith('.png')]
        for img in imagenes:
            img_path = os.path.join(TEMP_FOLDER, img)
            if os.path.isfile(img_path):
                pdf.set_font("Arial", 'B', 12)
                pdf.cell(0, 10, f"Gráfico: {img}", ln=True)
                pdf.image(img_path, w=160)
                pdf.ln(10)

    # Guardar
    os.makedirs(EXPORT_FOLDER, exist_ok=True)
    filename = f"resultados_exportados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    output_path = os.path.join(EXPORT_FOLDER, filename)
    pdf.output(output_path)
    return output_path

def generate_excel(datos, items):
    wb = Workbook()
    ws = wb.active
    ws.title = "Predicciones"

    # Tabla de predicciones
    if 'datos_predicciones' in datos:
        tabla = pd.DataFrame(datos['datos_predicciones'])
        ws.append(tabla.columns.tolist())
        for _, row in tabla.iterrows():
            ws.append(row.tolist())

    # Métricas
    if 'metricas' in datos:
        ws_metrics = wb.create_sheet("Métricas")
        for k, v in datos['metricas'].items():
            ws_metrics.append([k, v])

    # Coeficientes
    if 'modelo_coeficientes' in datos:
        ws_coef = wb.create_sheet("Coeficientes")
        ws_coef.append(['Variable', 'Coeficiente'])
        for k, v in datos['modelo_coeficientes'].items():
            ws_coef.append([k, round(v, 4)])

    # Imágenes desde static/temp
    if os.path.exists(TEMP_FOLDER):
        imagenes = [f for f in os.listdir(TEMP_FOLDER) if f.endswith('.png')]
        for img in imagenes:
            img_path = os.path.join(TEMP_FOLDER, img)
            if os.path.isfile(img_path):
                nombre_hoja = f"Graf_{img[:25]}"
                ws_img = wb.create_sheet(title=nombre_hoja)
                xl_img = XLImage(img_path)
                xl_img.width = 640
                xl_img.height = 480
                ws_img.add_image(xl_img, "A1")

    # Guardar archivo
    os.makedirs(EXPORT_FOLDER, exist_ok=True)
    filename = f"resultados_exportados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = os.path.join(EXPORT_FOLDER, filename)
    wb.save(output_path)
    return output_path
